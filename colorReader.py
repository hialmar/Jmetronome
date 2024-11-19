import traceback
import openpyxl
import argparse

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", type = str, help = "The name of the file to be read.")
    parser.add_argument("sheetNumber", type = str, help = "The number of the sheet in the file.")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.filename, data_only=True, read_only=True)
    fs = wb.worksheets[int(args.sheetNumber)]
    fs_count_row = fs.max_row
    if not fs_count_row:
        fs_count_row = 1000
    fs_count_col = fs.max_column
    if not fs_count_col:
        fs_count_col = 1000

    print("Tapez des coordonnées sous la forme x y ou FIN pour arrêter")

    nonfini = True;

    while nonfini :
        ch = input()

        if ch == "FIN":
            nonfini = False
            break

        tab = ch.split(' ')
        try:
            row = int(tab[0])
            column = int(tab[1])

            if row in range(1, fs_count_row + 1):
                if column in range(1, fs_count_col + 1):
                    cell_color = fs.cell(column=column, row=row)
                    bgColor = cell_color.fill.bgColor.index
                    fgColor = cell_color.fill.fgColor.index
                    if (bgColor == '00000000') or (bgColor == 'FF000000') or (bgColor == 64) :
                        if (fgColor == '00000000') or (fgColor == 'FF000000') or (fgColor == 64):
                            print (0)
                        else:
                            print (fgColor)
                    else:
                        print (bgColor)
                else:
                    print("ERR col incorrecte")
            else:
                print("ERR row incorrecte")

        except Exception:
            print(traceback.format_exc())
            print("ERR Il faut taper deux valeurs entières ou FIN")


